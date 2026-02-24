from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Avg, Count
from django.db.models.functions import TruncWeek
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from datetime import datetime, date, timedelta 
from django.utils.timezone import now 
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from .models import Vehiculo, Conductor, RegistroSalida, Ruta
from .forms import VehiculoForm, ConductorForm, RegistroGuardiaForm, EdicionAdminForm, RutaForm

# --- 1. SEMÁFORO ---
@login_required
def transporte_home(request):
    if request.user.groups.filter(name='Guardias').exists() and not request.user.is_superuser:
        return redirect('control_salida')
    return redirect('dashboard_transporte')

# --- 2. VISTA GUARDIA (Y ADMIN): Control de Salida ---
@login_required
def registro_control_salida(request):
    es_guardia = request.user.groups.filter(name='Guardias').exists()
    es_admin = request.user.is_superuser
    
    if not es_guardia and not es_admin:
         messages.error(request, "Acceso no autorizado.")
         return redirect('dashboard_transporte')

    if request.method == 'POST':
        form = RegistroGuardiaForm(request.POST)
        if form.is_valid():
            registro = form.save(commit=False)
            registro.registrado_por = request.user
            registro.save()
            messages.success(request, f'✅ Registro guardado: {registro.vehiculo.patente}')
            
            # NUEVO: Requerimiento 2.1 (Salidas múltiples)
            if form.cleaned_data.get('salidas_multiples'):
                # Redirige con parámetro para que el frontend mantenga el foco o limpie ciertos campos
                return redirect('control_salida') 
            return redirect('control_salida')
    else:
        form = RegistroGuardiaForm()

    # NUEVO: Requerimiento 2.2 (Orden y Paginación igual a la admin)
    registros_totales = RegistroSalida.objects.order_by('-fecha_registro')
    
    items_por_pagina = request.GET.get('page_size', '10')
    if items_por_pagina not in ['10', '20', '50', '100']:
        items_por_pagina = '10'

    paginator = Paginator(registros_totales, int(items_por_pagina))
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    custom_range = paginator.get_elided_page_range(page_obj.number, on_each_side=1, on_ends=1)
    
    return render(request, 'transporte/control_salida.html', {
        'form': form, 
        'page_obj': page_obj,
        'custom_range': custom_range,
        'page_size': items_por_pagina,
        'es_admin': es_admin
    })

# --- 3. VISTA ADMIN: Dashboard Completo ---
@login_required
def dashboard_transporte(request):
    if not request.user.is_staff and not request.user.is_superuser:
        return redirect('transporte_home')

    inicio_get = request.GET.get('inicio')
    fin_get = request.GET.get('fin')

    if inicio_get: request.session['transporte_inicio'] = inicio_get
    if fin_get: request.session['transporte_fin'] = fin_get

    fecha_hoy = now().date()
    inicio_defecto = date(fecha_hoy.year, fecha_hoy.month, 1).strftime('%Y-%m-%d')
    fin_defecto = fecha_hoy.strftime('%Y-%m-%d')

    fecha_inicio = request.session.get('transporte_inicio', inicio_defecto)
    fecha_fin = request.session.get('transporte_fin', fin_defecto)

    items_por_pagina = request.GET.get('page_size', '10')
    if items_por_pagina not in ['10', '20', '50', '100']:
        items_por_pagina = '10'
    
    registros_tabla = RegistroSalida.objects.filter(
        fecha_registro__date__range=[fecha_inicio, fecha_fin]
    ).select_related('vehiculo', 'ruta').order_by('-fecha_registro')[:3000]

    paginator = Paginator(registros_tabla, int(items_por_pagina))
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    custom_range = paginator.get_elided_page_range(page_obj.number, on_each_side=1, on_ends=1)

    return render(request, 'transporte/dashboard_admin.html', {
        'page_obj': page_obj,
        'page_size': items_por_pagina,
        'custom_range': custom_range,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    })

# --- 4. Edición (Admin) ---
@login_required
def editar_registro(request, registro_id):
    if not request.user.is_superuser:
        return redirect('dashboard_transporte')
        
    registro = get_object_or_404(RegistroSalida, id=registro_id)
    if request.method == 'POST':
        form = EdicionAdminForm(request.POST, instance=registro)
        if form.is_valid():
            form.save()
            messages.success(request, 'Registro actualizado.')
            return redirect('dashboard_transporte')
    else:
        form = EdicionAdminForm(instance=registro)
        
    return render(request, 'transporte/editar_registro.html', {'form': form, 'registro': registro})

# --- 5. UTILIDADES Y CREACIÓN ---
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@login_required
def exportar_excel_transporte(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Transporte_Aurora_{datetime.now().strftime("%d%m%Y")}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bitácora Transporte"
    
    # --- ESTILOS PROFESIONALES ---
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Azul corporativo
    header_font = Font(color='FFFFFF', bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = [
        'ID', 'Fecha y Hora', 'Movimiento', 'Patente', 'Tipo Vehículo', 'Capacidad',
        'Conductor', 'Empresa Externa', 'Ruta Base', 'Origen Real', 'Destino Real', 
        'Paradas Intermedias', 'Pax', '% Ocupación', 'Costo Total', 'Costo / Persona', 'Guardia'
    ]
    ws.append(headers)
    
    # Aplicar ancho y estilo a cabeceras
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = 18

    # --- DATOS ---
    registros = RegistroSalida.objects.select_related('vehiculo', 'ruta', 'conductor', 'registrado_por').order_by('-fecha_registro', '-id')
    
    for r in registros:
        origen_real = r.ruta.destino if r.tipo_movimiento == 'ENTRADA' else r.ruta.origen
        destino_real = r.ruta.origen if r.tipo_movimiento == 'ENTRADA' else r.ruta.destino
        empresa = r.conductor.nombre_empresa_externa if r.conductor.empresa_externa else 'Interno'
        
        cpp = round(r.valor_viaje / r.cantidad_pasajeros) if r.cantidad_pasajeros > 0 else 0
        ocupacion = round((r.cantidad_pasajeros / r.vehiculo.capacidad) * 100) if r.vehiculo.capacidad > 0 else 0

        row = [
            r.id, r.fecha_registro.strftime("%d/%m/%Y %H:%M"), r.get_tipo_movimiento_display(),
            r.vehiculo.patente, r.vehiculo.get_tipo_display(), r.vehiculo.capacidad,
            r.conductor.nombre, empresa, r.ruta.nombre, origen_real, destino_real,
            r.paradas_intermedias or 'Directo (Sin paradas)', r.cantidad_pasajeros,
            f"{ocupacion}%", f"${r.valor_viaje}", f"${cpp}",
            r.registrado_por.username if r.registrado_por else "Sistema"
        ]
        ws.append(row)
        
        # Aplicar bordes y centrado a filas de datos
        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.border = thin_border
            if col_num in [1, 3, 6, 13, 14, 15, 16]: 
                cell.alignment = center_align

    wb.save(response)
    return response
@login_required
def crear_vehiculo(request):
    if request.method == 'POST':
        form = VehiculoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Vehículo registrado.')
            return redirect('crear_vehiculo')
    else:
        form = VehiculoForm()
    
    vehiculos_existentes = Vehiculo.objects.filter(activo=True).order_by('patente')
    return render(request, 'transporte/crear_vehiculo.html', {'form': form, 'lista_vehiculos': vehiculos_existentes})

@login_required
def crear_conductor(request):
    if request.method == 'POST':
        form = ConductorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Conductor registrado.')
            return redirect('crear_conductor')
    else:
        form = ConductorForm()

    conductores_existentes = Conductor.objects.filter(activo=True).order_by('nombre')
    return render(request, 'transporte/crear_conductor.html', {'form': form, 'lista_conductores': conductores_existentes})

@login_required
def gestion_rutas(request):
    if not request.user.is_staff and not request.user.is_superuser:
        return redirect('transporte_home')
        
    if request.method == 'POST':
        form = RutaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Ruta creada.')
            return redirect('gestion_rutas')
    else:
        form = RutaForm()
    
    rutas = Ruta.objects.filter(activo=True)
    return render(request, 'transporte/gestion_rutas.html', {'form': form, 'rutas': rutas})

# --- NUEVAS FUNCIONES PARA DESHABILITAR (Requerimiento 1.1) ---
@login_required
def deshabilitar_vehiculo(request, vehiculo_id):
    if not request.user.is_superuser: return redirect('transporte_home')
    vehiculo = get_object_or_404(Vehiculo, id=vehiculo_id)
    vehiculo.activo = False
    vehiculo.save()
    messages.success(request, f'Vehículo {vehiculo.patente} deshabilitado.')
    return redirect('crear_vehiculo')

@login_required
def deshabilitar_conductor(request, conductor_id):
    if not request.user.is_superuser: return redirect('transporte_home')
    conductor = get_object_or_404(Conductor, id=conductor_id)
    conductor.activo = False
    conductor.save()
    messages.success(request, f'Conductor {conductor.nombre} deshabilitado.')
    return redirect('crear_conductor')

@login_required
def deshabilitar_ruta(request, ruta_id):
    if not request.user.is_superuser: return redirect('transporte_home')
    ruta = get_object_or_404(Ruta, id=ruta_id)
    ruta.activo = False
    ruta.save()
    messages.success(request, f'Ruta {ruta.nombre} deshabilitada.')
    return redirect('gestion_rutas')

# --- 6. API DASHBOARD ---
@login_required
def api_datos_dashboard(request):
    if not request.user.is_superuser: return JsonResponse({'error': '403'}, status=403)

    inicio = request.GET.get('inicio')
    fin = request.GET.get('fin')
    registros = RegistroSalida.objects.all()
    if inicio and fin: registros = registros.filter(fecha_registro__date__range=[inicio, fin])

    curva_raw = registros.annotate(semana=TruncWeek('fecha_registro')).values('semana').annotate(total=Sum('valor_viaje')).order_by('semana')
    curva_costos = [{'mes': f"Semana {x['semana'].strftime('%d/%m')}", 'total': x['total']} for x in curva_raw if x['semana']]

    # NUEVO: Requerimiento 1.4 (Se eliminaron las consultas y agrupaciones por patente)
    ocupacion_tipo = list(registros.values('vehiculo__tipo').annotate(avg_ocupacion=Avg('ocupacion_porcentaje')))
    costo_tipo = list(registros.values('vehiculo__tipo').annotate(total_costo=Sum('valor_viaje')))
    
    tot_costo = registros.aggregate(Sum('valor_viaje'))['valor_viaje__sum'] or 0
    tot_pax = registros.aggregate(Sum('cantidad_pasajeros'))['cantidad_pasajeros__sum'] or 1
    cpp_gen = tot_costo / tot_pax

    cpp_tipo = [{'label': i['vehiculo__tipo'], 'value': round(i['c']/i['p'] if i['p']>0 else 0)} 
                for i in registros.values('vehiculo__tipo').annotate(c=Sum('valor_viaje'), p=Sum('cantidad_pasajeros'))]

    costo_ruta = list(registros.values('ruta__nombre').annotate(total=Sum('valor_viaje')).order_by('-total'))

    cpp_ruta_raw = registros.values('ruta__nombre').annotate(c=Sum('valor_viaje'), p=Sum('cantidad_pasajeros'))
    cpp_ruta = [{'label': i['ruta__nombre'], 'value': round(i['c']/i['p'] if i['p']>0 else 0)} for i in cpp_ruta_raw]
    cpp_ruta.sort(key=lambda x: x['value'], reverse=True)

    return JsonResponse({
        'curva_costos': curva_costos,
        'ocupacion_tipo': ocupacion_tipo,
        'costo_tipo': costo_tipo,
        'cpp_general': round(cpp_gen),
        'cpp_tipo': cpp_tipo,
        'costo_ruta': costo_ruta,
        'cpp_ruta': cpp_ruta,
    })