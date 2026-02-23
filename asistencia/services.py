"""
Servicio de importación del Reporte de Estadía (Asistencia).

Estructura del archivo:
- Filas 0-8: Metadata del reporte (se ignoran)
- Fila 9  : Headers de columnas
- Fila 10+: Marcajes individuales (una fila = un marcaje)

Lógica:
  Por cada persona+fecha se toma:
    - hora_entrada = primer marcaje "Entrada"
    - hora_salida  = último marcaje "Salida"
"""
import openpyxl
from datetime import datetime, date, time
from collections import defaultdict
from django.db import transaction

from dotacion.models import Colaborador
from .models import RegistroAsistencia, Anomalia


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def _limpiar_rut(rut_raw):
    if not rut_raw:
        return None
    rut = str(rut_raw).replace('.', '').strip().upper()
    return rut if rut not in ('', 'NONE', 'NAN') else None


def _parse_fecha(valor):
    if not valor:
        return None
    if isinstance(valor, (date, datetime)):
        return valor.date() if isinstance(valor, datetime) else valor
    s = str(valor).strip()
    if s in ('', 'None', 'nan'):
        return None
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_hora(valor):
    if not valor:
        return None
    if isinstance(valor, time):
        return valor
    if isinstance(valor, datetime):
        return valor.time()
    s = str(valor).strip()
    for fmt in ('%H:%M:%S', '%H:%M'):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


# ─────────────────────────────────────────────
# Función principal
# ─────────────────────────────────────────────

def procesar_estadia(archivo_file, archivo_origen=None):
    """
    Lee el Reporte de Estadía y actualiza RegistroAsistencia.

    Returns:
        dict: registros_creados, registros_actualizados,
              ruts_no_encontrados, anomalias_creadas, errores
    """
    wb = openpyxl.load_workbook(archivo_file, read_only=True, data_only=True)
    ws = wb.active

    HEADER_ROW = 10
    DATA_START  = 11

    # Mapear headers
    headers = {}
    for col_idx, cell in enumerate(ws[HEADER_ROW], start=1):
        val = str(cell.value or '').strip().upper()
        if val:
            headers[val] = col_idx

    for campo in ['RUT', 'FECHA', 'HORA', 'MOVIMIENTO']:
        if campo not in headers:
            raise ValueError(
                f"El archivo no tiene la columna obligatoria '{campo}'. "
                "Verifica que sea el Reporte de Estadía correcto."
            )

    def col(row, nombre):
        idx = headers.get(nombre)
        return row[idx - 1].value if idx else None

    # Agrupar marcajes por (rut, fecha)
    marcajes = defaultdict(lambda: {'entradas': [], 'salidas': []})

    for row in ws.iter_rows(min_row=DATA_START, values_only=False):
        rut   = _limpiar_rut(col(row, 'RUT'))
        fecha = _parse_fecha(col(row, 'FECHA'))
        hora  = _parse_hora(col(row, 'HORA'))
        movim = str(col(row, 'MOVIMIENTO') or '').strip().capitalize()

        if not all([rut, fecha, hora]):
            continue

        if movim == 'Entrada':
            marcajes[(rut, fecha)]['entradas'].append(hora)
        elif movim == 'Salida':
            marcajes[(rut, fecha)]['salidas'].append(hora)

    wb.close()

    # Pre-cargar colaboradores en memoria
    ruts_en_reporte = {rut for rut, _ in marcajes.keys()}
    colaboradores_map = {
        c.rut: c
        for c in Colaborador.objects.filter(rut__in=ruts_en_reporte)
    }

    registros_creados     = 0
    registros_actualizados = 0
    ruts_desconocidos     = set()
    anomalias_creadas     = 0
    errores               = []

    with transaction.atomic():
        for (rut, fecha), tiempos in marcajes.items():

            colaborador = colaboradores_map.get(rut)
            if not colaborador:
                ruts_desconocidos.add(rut)
                continue

            entradas = sorted(tiempos['entradas'])
            salidas  = sorted(tiempos['salidas'])

            hora_entrada = entradas[0]  if entradas else None
            hora_salida  = salidas[-1]  if salidas  else None

            try:
                registro, created = RegistroAsistencia.objects.update_or_create(
                    colaborador=colaborador,
                    fecha=fecha,
                    defaults={
                        'hora_entrada'   : hora_entrada,
                        'hora_salida'    : hora_salida,
                        'archivo_origen' : archivo_origen,
                    }
                )

                if created:
                    registros_creados += 1
                else:
                    registros_actualizados += 1

                # Recalcular anomalías
                registro.anomalias.all().delete()

                if hora_entrada and not hora_salida:
                    Anomalia.objects.create(
                        registro=registro,
                        tipo='SIN_MARCA',
                        observacion='Solo tiene marca de entrada, falta salida.'
                    )
                    anomalias_creadas += 1
                elif not hora_entrada and hora_salida:
                    Anomalia.objects.create(
                        registro=registro,
                        tipo='SIN_MARCA',
                        observacion='Solo tiene marca de salida, falta entrada.'
                    )
                    anomalias_creadas += 1

            except Exception as e:
                errores.append(f"{rut} {fecha}: {str(e)}")

    return {
        'registros_creados'     : registros_creados,
        'registros_actualizados': registros_actualizados,
        'ruts_no_encontrados'   : len(ruts_desconocidos),
        'anomalias_creadas'     : anomalias_creadas,
        'errores'               : errores,
    }