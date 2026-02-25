"""
Servicio de importación del Reporte de Fichas (Dotación).

Estructura del archivo:
- Filas 0-8: Metadata y encabezados del reporte (se ignoran)
- Fila 9  : Headers de columnas (RUT, NOMBRES, ÁREA, etc.)
- Fila 10+: Datos de colaboradores
"""
import openpyxl
from datetime import datetime, date
from django.db import transaction

from .models import Colaborador


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────

def _limpiar_str(valor):
    if valor is None:
        return ''
    return str(valor).strip()


def _limpiar_rut(rut_raw):
    if not rut_raw:
        return None
    rut = str(rut_raw).replace('.', '').strip().upper()
    return rut if rut not in ('', 'NONE', 'NAN') else None


def _parse_fecha(valor):
    if not valor:
        return None
    if isinstance(valor, (date, datetime)):
        return valor.date() if isinstance(valor, datetime) else valor
    s = str(valor).strip()
    if s in ('', '00-00-0000', '00/00/0000', 'None', 'nan'):
        return None
    for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ─────────────────────────────────────────────
# Función principal
# ─────────────────────────────────────────────

def procesar_fichas(archivo_file):
    """
    Lee el Reporte de Fichas y hace upsert en Colaborador.

    Returns:
        dict con claves: creados, actualizados, omitidos, errores (lista)
    """
    wb = openpyxl.load_workbook(archivo_file, read_only=True, data_only=True)
    ws = wb.active

    HEADER_ROW = 10
    DATA_START  = 11

    # Mapear headers por nombre
    headers = {}
    for col_idx, cell in enumerate(ws[HEADER_ROW], start=1):
        val = _limpiar_str(cell.value).upper()
        if val:
            headers[val] = col_idx

    # Verificación mínima
    for campo in ['RUT', 'NOMBRES', 'PRIMER APELLIDO']:
        if campo not in headers:
            raise ValueError(
                f"El archivo no tiene la columna obligatoria '{campo}'. "
                "Verifica que sea el Reporte de Fichas correcto."
            )

    def col(row, nombre):
        idx = headers.get(nombre)
        if idx is None:
            return None
        return row[idx - 1].value

    creados     = 0
    actualizados = 0
    omitidos    = 0
    errores     = []

    with transaction.atomic():
        for row_num, row in enumerate(
            ws.iter_rows(min_row=DATA_START, values_only=False), start=DATA_START
        ):
            rut = _limpiar_rut(row[headers['RUT'] - 1].value)

            if not rut:
                omitidos += 1
                continue

            nombres  = _limpiar_str(col(row, 'NOMBRES'))
            paterno  = _limpiar_str(col(row, 'PRIMER APELLIDO'))
            materno  = _limpiar_str(col(row, 'SEGUNDO APELLIDO'))
            nombre_completo = f"{nombres} {paterno} {materno}".strip()

            estado_ficha = _limpiar_str(col(row, 'ESTADO FICHA'))
            estado_lower = estado_ficha.lower()
            if estado_lower == 'vigente':
                estado = 'VIGENTE'
            elif estado_lower in ('finiquitado', 'finiquito'):
                estado = 'FINIQUITADO'
            else:
                estado = 'FINIQUITADO'   # cualquier otro estado de GREX = inactivo

            estado_rec = _limpiar_str(col(row, 'ESTADO RECOMENDABLE')).upper()
            es_recomendable = estado_rec != 'NO RECOMENDABLE'

            try:
                defaults = {
                    'nombre_completo'       : nombre_completo,
                    'codigo_ficha'          : col(row, 'CÓDIGO FICHA'),
                    'cargo'                 : _limpiar_str(col(row, 'CARGO')) or None,
                    'centro_costo'          : _limpiar_str(col(row, 'CENTRO COSTO')) or None,
                    'area'                  : _limpiar_str(col(row, 'ÁREA')) or None,
                    'seccion'               : _limpiar_str(col(row, 'SECCIÓN')) or None,
                    'turno'                 : _limpiar_str(col(row, 'TURNO')) or None,
                    'tipo_contrato'         : _limpiar_str(col(row, 'TIPO CONTRATO')) or None,
                    'estado_ficha'          : estado_ficha or None,
                    'estado_civil'          : _limpiar_str(col(row, 'ESTADO CIVIL')) or None,
                    'fecha_ingreso'         : _parse_fecha(col(row, 'FECHA INGRESO')),
                    'fecha_termino_contrato': _parse_fecha(col(row, 'FECHA TÉRMINO CONTRATO')),
                    'fecha_nacimiento'      : _parse_fecha(col(row, 'FECHA NACIMIENTO')),
                    'sexo'                  : _limpiar_str(col(row, 'SEXO')) or None,
                    'nacionalidad'          : _limpiar_str(col(row, 'NACIONALIDAD')) or None,
                    'comuna'                : _limpiar_str(col(row, 'COMUNA')) or None,
                    'ciudad'                : _limpiar_str(col(row, 'CIUDAD')) or None,
                    'direccion'             : _limpiar_str(col(row, 'DIRECCIÓN')) or None,
                    'escolaridad'           : _limpiar_str(col(row, 'ESCOLARIDAD')) or None,
                    'email'                 : _limpiar_str(col(row, 'EMAIL')) or None,
                    'telefono'              : _limpiar_str(col(row, 'TELÉFONO')) or None,
                    'es_recomendable'       : es_recomendable,
                    'estado'     : estado,
                    'estado_ficha': estado_ficha or None,
                }
                obj, created = Colaborador.objects.update_or_create(
                    rut=rut,
                    defaults=defaults,
                )
                if created:
                    creados += 1
                else:
                    actualizados += 1

            except Exception as e:
                errores.append(f"Fila {row_num}: {rut} → {str(e)}")

    wb.close()

    return {
        'creados'     : creados,
        'actualizados': actualizados,
        'omitidos'    : omitidos,
        'errores'     : errores,
    }