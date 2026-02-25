from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.db.models.functions import TruncWeek
from datetime import date, timedelta
from .forms import CargaFichasForm
from .models import Colaborador
from .services import procesar_fichas


@login_required
def index(request):
    if request.method == 'POST':
        form = CargaFichasForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                resultado = procesar_fichas(request.FILES['archivo'])
                messages.success(
                    request,
                    f"✅ Carga completa. "
                    f"Nuevos: {resultado['creados']} | "
                    f"Actualizados: {resultado['actualizados']} | "
                    f"Omitidos: {resultado['omitidos']}"
                )
                if resultado['errores']:
                    messages.warning(request, f"⚠️ {len(resultado['errores'])} filas con error.")
                    for err in resultado['errores'][:5]:
                        messages.warning(request, err)
            except ValueError as e:
                messages.error(request, f"❌ {str(e)}")
            except Exception as e:
                messages.error(request, f"❌ Error inesperado: {str(e)}")
            return redirect('dotacion_index')
    else:
        form = CargaFichasForm()

    hoy = date.today()
    return render(request, 'dotacion/index.html', {
        'form'                : form,
        'fecha_inicio_default': date(hoy.year, 1, 1).strftime('%Y-%m-%d'),
        'fecha_fin_default'   : hoy.strftime('%Y-%m-%d'),
    })


@login_required
def api_kpis(request):
    inicio_str = request.GET.get('inicio')
    fin_str    = request.GET.get('fin')
    hoy        = date.today()

    try:
        from datetime import datetime
        inicio = datetime.strptime(inicio_str, '%Y-%m-%d').date() if inicio_str else date(hoy.year, 1, 1)
        fin    = datetime.strptime(fin_str,    '%Y-%m-%d').date() if fin_str    else hoy
    except ValueError:
        inicio = date(hoy.year, 1, 1)
        fin    = hoy

    vigentes = Colaborador.objects.filter(estado='VIGENTE', fecha_ingreso__lte=fin)

    total_vigentes  = vigentes.count()
    total_historico = Colaborador.objects.count()
    sin_contacto    = vigentes.filter(
        Q(email__isnull=True) | Q(email='') |
        Q(telefono__isnull=True) | Q(telefono='')
    ).count()

    # 1. Dotación activa por semana
    todos = list(
        Colaborador.objects
        .filter(fecha_ingreso__isnull=False, fecha_ingreso__lte=fin)
        .values('fecha_ingreso', 'fecha_termino_contrato')
    )
    semana = inicio - timedelta(days=inicio.weekday())
    dotacion_labels, dotacion_values = [], []
    while semana <= fin:
        count = sum(
            1 for c in todos
            if c['fecha_ingreso'] <= semana
            and (c['fecha_termino_contrato'] is None or c['fecha_termino_contrato'] >= semana)
        )
        dotacion_labels.append(semana.strftime('%d/%m/%Y'))
        dotacion_values.append(count)
        semana += timedelta(weeks=1)

    # 2. Contrataciones por semana
    contrataciones_raw = (
        Colaborador.objects
        .filter(fecha_ingreso__isnull=False, fecha_ingreso__range=[inicio, fin])
        .annotate(semana=TruncWeek('fecha_ingreso'))
        .values('semana').annotate(total=Count('rut')).order_by('semana')
    )
    contrataciones_labels = [r['semana'].strftime('%d/%m/%Y') for r in contrataciones_raw]
    contrataciones_values = [r['total'] for r in contrataciones_raw]

    # 3. Rango etario
    rangos = {'18-25': 0, '26-35': 0, '36-45': 0, '46-55': 0, '56+': 0}
    for c in vigentes.filter(fecha_nacimiento__isnull=False):
        edad = c.edad
        if edad is None: continue
        if   18 <= edad <= 25: rangos['18-25'] += 1
        elif 26 <= edad <= 35: rangos['26-35'] += 1
        elif 36 <= edad <= 45: rangos['36-45'] += 1
        elif 46 <= edad <= 55: rangos['46-55'] += 1
        elif edad >= 56:       rangos['56+']   += 1

    # 4. Nivel educacional
    escolaridad_data = list(
        vigentes.exclude(escolaridad__isnull=True).exclude(escolaridad='')
        .values('escolaridad').annotate(total=Count('rut')).order_by('-total')
    )

    # 5. Nacionalidades
    nacionalidad_data = list(
        vigentes.exclude(nacionalidad__isnull=True).exclude(nacionalidad='')
        .values('nacionalidad').annotate(total=Count('rut')).order_by('-total')[:8]
    )

    # 6. Comunas
    comuna_data = list(
        vigentes.exclude(comuna__isnull=True).exclude(comuna='')
        .values('comuna').annotate(total=Count('rut')).order_by('-total')[:12]
    )

    # 7. Género
    sexo_data = list(
        vigentes.exclude(sexo__isnull=True).exclude(sexo='')
        .values('sexo').annotate(total=Count('rut')).order_by('-total')
    )

    # 8. Permanencia
    rangos_perm = {
        '< 3 meses': 0, '3-6 meses': 0, '6-12 meses': 0,
        '1-2 años': 0, '2-5 años': 0, '5+ años': 0
    }
    for c in vigentes.filter(fecha_ingreso__isnull=False):
        m = c.meses_permanencia
        if m is None: continue
        if   m < 3:  rangos_perm['< 3 meses']  += 1
        elif m < 6:  rangos_perm['3-6 meses']   += 1
        elif m < 12: rangos_perm['6-12 meses']  += 1
        elif m < 24: rangos_perm['1-2 años']     += 1
        elif m < 60: rangos_perm['2-5 años']     += 1
        else:        rangos_perm['5+ años']      += 1

    # 9. Tipo contrato
    planta_data = list(
        vigentes.exclude(tipo_contrato__isnull=True).exclude(tipo_contrato='')
        .values('tipo_contrato').annotate(total=Count('rut')).order_by('-total')
    )

    return JsonResponse({
        'kpi_vigentes'    : total_vigentes,
        'kpi_historico'   : total_historico,
        'kpi_sin_contacto': sin_contacto,
        'dotacion'        : {'labels': dotacion_labels,        'values': dotacion_values},
        'contrataciones'  : {'labels': contrataciones_labels,  'values': contrataciones_values},
        'edad'            : {'labels': list(rangos.keys()),     'values': list(rangos.values())},
        'escolaridad'     : {'labels': [d['escolaridad']   for d in escolaridad_data],  'values': [d['total'] for d in escolaridad_data]},
        'nacionalidad'    : {'labels': [d['nacionalidad']  for d in nacionalidad_data], 'values': [d['total'] for d in nacionalidad_data]},
        'comuna'          : {'labels': [d['comuna']        for d in comuna_data],       'values': [d['total'] for d in comuna_data]},
        'sexo'            : {'labels': [d['sexo']          for d in sexo_data],         'values': [d['total'] for d in sexo_data]},
        'permanencia'     : {'labels': list(rangos_perm.keys()), 'values': list(rangos_perm.values())},
        'tipo_contrato'   : {'labels': [d['tipo_contrato'] for d in planta_data],       'values': [d['total'] for d in planta_data]},
    })


@login_required
def api_buscar(request):
    if not request.user.is_staff and not request.user.is_superuser:
        return JsonResponse({'error': '403'}, status=403)

    q = request.GET.get('q', '').strip()
    if len(q) < 3:
        return JsonResponse({'resultados': []})

    colaboradores = (
        Colaborador.objects
        .filter(Q(rut__icontains=q) | Q(nombre_completo__icontains=q))
        .values('rut', 'nombre_completo', 'cargo', 'centro_costo', 'estado', 'motivo_bloqueo')
        [:10]
    )
    return JsonResponse({'resultados': list(colaboradores)})


@login_required
def lista_bloqueados(request):
    if not request.user.is_staff and not request.user.is_superuser:
        return redirect('dotacion_index')

    from .models import PersonaBloqueada

    if request.method == 'POST':
        rut    = request.POST.get('rut', '').strip()
        nombre = request.POST.get('nombre', '').strip().upper()
        motivo = request.POST.get('motivo', '').strip().upper()
        foto   = request.FILES.get('foto')

        if rut and nombre and motivo:
            rut_limpio = rut.replace('.', '').replace('-', '').upper()
            try:
                cuerpo, dv = rut_limpio[:-1], rut_limpio[-1]
                rut = f"{int(cuerpo):,}".replace(',', '.') + '-' + dv
            except ValueError:
                pass

            PersonaBloqueada.objects.update_or_create(
                rut      = rut,
                defaults = {
                    'nombre_completo': nombre,
                    'motivo'         : motivo,
                    'foto'           : foto,
                    'bloqueado_por'  : request.user,
                    'activo'         : True,
                }
            )
            messages.success(request, f'✅ {nombre} agregado a la lista de bloqueados.')
            return redirect('dotacion_bloqueados')

    bloqueados = PersonaBloqueada.objects.filter(activo=True)
    return render(request, 'dotacion/lista_bloqueados.html', {'bloqueados': bloqueados})


@login_required
def desbloquear_persona(request, pk):
    if not request.user.is_staff and not request.user.is_superuser:
        return JsonResponse({'ok': False}, status=403)

    from .models import PersonaBloqueada
    persona        = get_object_or_404(PersonaBloqueada, pk=pk)
    persona.activo = False
    persona.save()

    return JsonResponse({'ok': True, 'nombre': persona.nombre_completo})


@login_required
def descargar_plantilla_bloqueados(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bloqueados"

    headers     = ['RUT', 'NOMBRE_COMPLETO', 'MOTIVO']
    header_fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

    ejemplos = [
        ['12.345.678-9', 'GARCIA LOPEZ JUAN', 'Conducta inapropiada en instalaciones'],
        ['9.876.543-2',  'PEREZ SOTO MARIA',  'Robo comprobado'],
    ]
    for row in ejemplos:
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=plantilla_bloqueados.xlsx'
    wb.save(response)
    return response


@login_required
def carga_masiva_bloqueados(request):
    if request.method != 'POST':
        return redirect('dotacion_bloqueados')

    if not request.user.is_staff and not request.user.is_superuser:
        return redirect('dotacion_bloqueados')

    from .models import PersonaBloqueada
    import openpyxl

    archivo = request.FILES.get('archivo_masivo')
    if not archivo:
        messages.error(request, '❌ No se recibió ningún archivo.')
        return redirect('dotacion_bloqueados')

    try:
        wb      = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
        ws      = wb.active
        creados = omitidos = 0
        errores = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            rut    = str(row[0]).strip() if row[0] is not None else ''
            nombre = str(row[1]).strip().upper() if row[1] is not None else ''
            motivo = str(row[2]).strip().upper() if row[2] is not None else ''

            if not rut or not nombre or not motivo:
                omitidos += 1
                if len(errores) < 10:
                    errores.append(
                        f"Fila {row_num}: rut='{rut}' | nombre='{nombre[:20]}' | motivo='{motivo[:20]}'"
                    )
                continue

            rut_limpio = rut.replace('.', '').replace('-', '').replace(' ', '').upper()
            try:
                cuerpo         = rut_limpio[:-1]
                dv             = rut_limpio[-1]
                rut_formateado = f"{int(cuerpo):,}".replace(',', '.') + '-' + dv
            except (ValueError, IndexError):
                errores.append(f"Fila {row_num}: RUT no procesable → '{rut}'")
                omitidos += 1
                continue

            PersonaBloqueada.objects.update_or_create(
                rut      = rut_formateado,
                defaults = {
                    'nombre_completo': nombre,
                    'motivo'         : motivo,
                    'bloqueado_por'  : request.user,
                    'activo'         : True,
                }
            )
            creados += 1

        wb.close()
        messages.success(request, f'✅ Registrados: {creados} | Omitidos: {omitidos}')
        if errores:
            messages.warning(request, '⚠️ Primeras omisiones:')
            for e in errores:
                messages.warning(request, e)

    except Exception as e:
        messages.error(request, f'❌ Error procesando archivo: {str(e)}')

    return redirect('dotacion_bloqueados')